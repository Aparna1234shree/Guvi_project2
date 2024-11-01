"""
utils/excel_functions.py

This module provides the ExcelFunctions class, which includes methods for reading from and writing to an
Excel spreadsheet using the OpenPyXL library.

Purpose:
- To encapsulate common Excel file operations, such as reading and writing data in a specified worksheet.
"""
from openpyxl import load_workbook

class ExcelFunctions:
    def __init__(self, file_name, sheet_name):
        """Initialize the ExcelFunctions with the given Excel file and sheet name.

                Args:
                    file_name (str): The name of the Excel file to operate on.
                    sheet_name (str): The name of the worksheet within the Excel file.
        """
        self.file = file_name
        self.sheet = sheet_name

    def read_data(self, row, col):
        """Read data from a specific cell in the Excel sheet.

               This method loads the workbook, accesses the specified sheet, and retrieves the value from the cell
               located at the provided row and column indices.

               Args:
                   row (int): The row number of the cell to read (1-indexed).
                   col (int): The column number of the cell to read (1-indexed).

               Returns:
                   The value of the specified cell, or None if the cell is empty.
        """
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row, column=col).value

    def write_data(self, row, col, data):
        """Write data to a specific cell in the Excel sheet.

                This method loads the workbook, accesses the specified sheet, writes the provided data to the cell
                located at the given row and column indices, and then saves the workbook.

                Args:
                    row (int): The row number of the cell to write to (1-indexed).
                    col (int): The column number of the cell to write to (1-indexed).
                    data: The data to be written to the specified cell.
        """
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row, column=col).value = data
        workbook.save(self.file)
